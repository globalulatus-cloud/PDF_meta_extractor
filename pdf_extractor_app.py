"""
EN→ES Activity Card Extractor
Extracts font-size-28 text anchored by ▲ ● ■ symbols from source (English)
and target (Spanish) PDFs, then exports matched pairs to Excel.

Compatible with PyMuPDF (fitz) 1.18+
"""

import io
import re
import fitz  # PyMuPDF
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Constants ────────────────────────────────────────────────────────────────
ANCHOR_SYMBOLS = {"▲": "Triangle", "●": "Circle", "■": "Square"}
FONT_SIZE_TARGET = 28          # pt – primary extraction target
FONT_SIZE_TOLERANCE = 2        # ±pt allowed
CARD_PATTERNS = [
    # English:  "Activity Card 1-1 ▲"  /  "Activity Card 1-1 ●"
    re.compile(
        r"Activity\s+Card\s+(\d+[-–]\d+)\s*([▲●■])",
        re.IGNORECASE
    ),
    # Spanish:  "Tarjeta de actividad 1-1 ▲"
    re.compile(
        r"Tarjeta\s+de\s+actividad\s+(\d+[-–]\d+)\s*([▲●■])",
        re.IGNORECASE
    ),
]
# Grade/Unit/Lesson footer patterns
FOOTER_EN = re.compile(
    r"Grade\s+(\d+)[,\s]+Unit\s+(\d+)[,\s]+Lesson\s+(\d+)",
    re.IGNORECASE
)
FOOTER_ES = re.compile(
    r"Grado\s+(\d+)[,\s]+Unidad\s+(\d+)[,\s]+Lecci[oó]n\s+(\d+)",
    re.IGNORECASE
)


# ── Extraction helpers ────────────────────────────────────────────────────────

def extract_pages(pdf_source,
                  font_size: float = FONT_SIZE_TARGET,
                  tolerance: float = FONT_SIZE_TOLERANCE,
                  progress_bar=None) -> list[dict]:
    """
    Memory-efficient extraction: streams the PDF page-by-page.
    pdf_source can be:
      - a file path (str / Path)  → opened directly from disk, lowest RAM use
      - raw bytes                 → loaded into a BytesIO buffer
    raw_blocks are NOT stored after extraction to keep RAM usage low.
    """
    import os
    if isinstance(pdf_source, (str, os.PathLike)):
        doc = fitz.open(str(pdf_source))
    else:
        doc = fitz.open(stream=pdf_source, filetype="pdf")

    total_pages = doc.page_count
    pages = []

    for page_num, page in enumerate(doc, start=1):
        # Collect all text spans with their sizes
        spans_by_size: dict[float, list[str]] = {}
        all_spans: list[dict] = []

        blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
        for block in blocks:
            if block.get("type") != 0:   # 0 = text block
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span["text"].strip()
                    if not text:
                        continue
                    size = round(span["size"], 1)
                    all_spans.append({"text": text, "size": size, "bbox": span["bbox"]})
                    spans_by_size.setdefault(size, []).append(text)

        # Full page text (for regex searches)
        full_text = page.get_text("text")

        # ── Identify card_id and anchor symbol ──────────────────────────────
        card_id, anchor = None, None
        for pat in CARD_PATTERNS:
            m = pat.search(full_text)
            if m:
                card_id = m.group(1)
                anchor  = m.group(2)
                break

        # Also scan spans directly (symbol may be separate span)
        if anchor is None:
            for sp in all_spans:
                for sym in ANCHOR_SYMBOLS:
                    if sym in sp["text"]:
                        anchor = sym
                        break
                if anchor:
                    break

        # ── All spans at target font size ────────────────────────────────────
        all_f28 = [
            sp["text"] for sp in all_spans
            if abs(sp["size"] - font_size) <= tolerance
        ]

        # ── Title: largest font on page (usually the lesson title) ──────────
        title = None
        skip_re = re.compile(
            r"copyright|\xa9|All rights reserved|Heinemann|Houghton|Publishing"
            r"|Grade \d|Unit \d|Lesson \d|Grado \d|Unidad \d|Lecci",
            re.IGNORECASE,
        )
        if all_spans:
            size_buckets = {}
            for sp in all_spans:
                txt = sp["text"]
                if len(txt) <= 3:
                    continue
                if any(c in txt for c in ANCHOR_SYMBOLS):
                    continue
                if skip_re.search(txt):
                    continue
                sz = sp["size"]
                size_buckets.setdefault(sz, []).append(txt)
            if size_buckets:
                best_size = max(size_buckets.keys())
                title = " ".join(size_buckets[best_size])

        # ── Numbered steps (lines starting with 1–9) ─────────────────────────
        step_pat = re.compile(r"^\s*[1-9]\s+(.+)", re.MULTILINE)
        steps = [m.group(1).strip() for m in step_pat.finditer(full_text)]

        # ── Footer ───────────────────────────────────────────────────────────
        footer = None
        for pat in (FOOTER_EN, FOOTER_ES):
            m = pat.search(full_text)
            if m:
                footer = m.group(0).strip()
                break

        pages.append({
            "page_num":    page_num,
            "card_id":     card_id,
            "anchor":      anchor,
            "anchor_name": ANCHOR_SYMBOLS.get(anchor) if anchor else None,
            "title":       title,
            "steps":       steps,
            "footer":      footer,
            "all_f28":     all_f28,
            # raw_blocks intentionally omitted — too large for 500 MB PDFs
            "raw_blocks":  all_spans if page_num == 1 else [],
        })

        # Update progress bar if provided (Streamlit)
        if progress_bar is not None:
            progress_bar.progress(page_num / total_pages)

        # Release page resources immediately
        page = None

    doc.close()
    return pages


def match_pages(src_pages: list[dict], tgt_pages: list[dict]) -> list[dict]:
    """
    Match source and target pages by (card_id, anchor).

    For unmatched target pages (ES-only cards with no card_id in EN):
      - Try page-proximity: find the nearest EN source page by page number
        whose anchor symbol matches, within a configurable window.
      - If still no match, mark as ES-Only (genuinely absent from EN).
    """
    PROXIMITY_WINDOW = 20   # max page-number distance to attempt proximity match

    def build_index(pages):
        idx = {}
        for p in pages:
            key = (p["card_id"], p["anchor"])
            if key[0]:
                idx[key] = p
        return idx

    src_idx = build_index(src_pages)
    tgt_idx = build_index(tgt_pages)

    # Build a list of unmatched source pages (no target found yet) for proximity
    src_by_pagenum = {p["page_num"]: p for p in src_pages}

    rows = []
    matched_src_keys  = set()
    used_src_pagenums = set()   # track src pages consumed by proximity matches

    # Pass 1: primary match by card_id + anchor
    for key, sp in src_idx.items():
        tp = tgt_idx.get(key)
        rows.append(_build_row(sp, tp, matched=bool(tp)))
        matched_src_keys.add(key)
        if tp:
            used_src_pagenums.add(sp["page_num"])

    # Pass 2: handle target pages with no card_id match
    unmatched_tgt_keys = [k for k in tgt_idx if k not in matched_src_keys]

    for key in unmatched_tgt_keys:
        tp = tgt_idx[key]
        tgt_pnum = tp["page_num"]
        tgt_anchor = tp["anchor"]

        # Proximity search: walk outward from the target page number
        # looking for a source page with the same anchor that hasn't been used
        best_src = None
        for delta in range(1, PROXIMITY_WINDOW + 1):
            for candidate_pnum in (tgt_pnum - delta, tgt_pnum + delta):
                sp = src_by_pagenum.get(candidate_pnum)
                if sp is None:
                    continue
                if sp["page_num"] in used_src_pagenums:
                    continue
                if sp["anchor"] != tgt_anchor:
                    continue
                # Also require that this source page has no card_id
                # (so we don't steal a properly-matched source page)
                if sp["card_id"] is not None:
                    continue
                best_src = sp
                break
            if best_src:
                break

        if best_src:
            used_src_pagenums.add(best_src["page_num"])
            rows.append(_build_row(best_src, tp, matched=True,
                                   match_note="Proximity match"))
        else:
            # Genuinely ES-only — no EN counterpart found
            rows.append(_build_row(None, tp, matched=False,
                                   match_note="ES-Only (no EN counterpart)"))

    # Fallback: if no card IDs found anywhere, zip by page order
    if not rows:
        for i, sp in enumerate(src_pages):
            tp = tgt_pages[i] if i < len(tgt_pages) else None
            rows.append(_build_row(sp, tp, matched=True))

    # Sort by source page, then target page
    rows.sort(key=lambda r: (r["Source Page"] or 9999, r["Target Page"] or 9999))
    return rows


def _build_row(sp, tp, matched: bool, match_note: str = "") -> dict:
    def pg(p):
        return p["page_num"] if p else ""

    def title(p):
        return p["title"] or "" if p else ""

    def card(p):
        cid  = p.get("card_id")  or ""
        anc  = p.get("anchor")   or ""
        anam = p.get("anchor_name") or ""
        return f"{cid} {anc}".strip(), anam

    def steps(p):
        return " | ".join(p["steps"]) if p and p["steps"] else ""

    def f28(p):
        return " | ".join(p["all_f28"]) if p and p["all_f28"] else ""

    def footer(p):
        return p.get("footer") or "" if p else ""

    src_card, src_aname = card(sp) if sp else ("", "")
    tgt_card, tgt_aname = card(tp) if tp else ("", "")

    return {
        "Source Page":         pg(sp),
        "Target Page":         pg(tp),
        "Card ID (Source)":    src_card,
        "Anchor Symbol":       (sp["anchor"] if sp and sp["anchor"] else
                                tp["anchor"] if tp and tp["anchor"] else ""),
        "Anchor Type":         src_aname or tgt_aname,
        "Source Title (EN)":   title(sp),
        "Target Title (ES)":   title(tp),
        "Source Steps (EN)":   steps(sp),
        "Target Steps (ES)":   steps(tp),
        "Source Font-28 (EN)": f28(sp),
        "Target Font-28 (ES)": f28(tp),
        "Source Footer":       footer(sp),
        "Target Footer":       footer(tp),
        "Matched":    "✓" if matched else "✗ (Unmatched)",
        "Match Note": match_note,
    }


# ── Excel export ──────────────────────────────────────────────────────────────

HEADER_FILL   = PatternFill("solid", fgColor="2E4057")   # dark navy
SRC_FILL      = PatternFill("solid", fgColor="D9EAD3")   # light green
TGT_FILL      = PatternFill("solid", fgColor="CFE2F3")   # light blue
META_FILL     = PatternFill("solid", fgColor="FFF2CC")   # light yellow
UNMATCH_FILL  = PatternFill("solid", fgColor="F4CCCC")   # light red
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT     = Font(name="Arial", size=9)
THIN_BORDER   = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

COLUMNS = [
    ("Source Page",         8),
    ("Target Page",         8),
    ("Card ID (Source)",    14),
    ("Anchor Symbol",       10),
    ("Anchor Type",         12),
    ("Source Title (EN)",   30),
    ("Target Title (ES)",   30),
    ("Source Steps (EN)",   45),
    ("Target Steps (ES)",   45),
    ("Source Font-28 (EN)", 40),
    ("Target Font-28 (ES)", 40),
    ("Source Footer",       22),
    ("Target Footer",       22),
    ("Matched",             14),
    ("Match Note",           22),
]

SRC_COLS = {"Source Page", "Source Title (EN)", "Source Steps (EN)",
            "Source Font-28 (EN)", "Source Footer", "Card ID (Source)"}
TGT_COLS = {"Target Page", "Target Title (ES)", "Target Steps (ES)",
            "Target Font-28 (ES)", "Target Footer"}
META_COLS = {"Anchor Symbol", "Anchor Type", "Matched", "Match Note"}


import re as _re
_ILLEGAL_CHARS = _re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f"      # C0 control chars (keep \t \n \r)
    r"\ud800-\udfff"                            # lone surrogates
    r"\ufffe\uffff]"                            # non-characters
)

def _sanitize(value) -> str:
    """Strip illegal XML/openpyxl characters from a cell value."""
    if not isinstance(value, str):
        return value
    cleaned = _ILLEGAL_CHARS.sub("", value)
    # Also replace vertical-tab / form-feed with space
    cleaned = cleaned.replace("\x0b", " ").replace("\x0c", " ")
    return cleaned


def build_excel(rows: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EN-ES Extraction"

    # Header row
    headers = [c[0] for c in COLUMNS]
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        unmatched = row.get("Matched", "✓") != "✓"
        for col_idx, (header, _) in enumerate(COLUMNS, start=1):
            value = _sanitize(row.get(header, ""))
            cell  = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = BODY_FONT
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            if unmatched:
                cell.fill = UNMATCH_FILL
            elif header in SRC_COLS:
                cell.fill = SRC_FILL
            elif header in TGT_COLS:
                cell.fill = TGT_FILL
            elif header in META_COLS:
                cell.fill = META_FILL

        ws.row_dimensions[row_idx].height = 55

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Extraction Summary"
    ws2["A1"].font = Font(name="Arial", bold=True, size=14)

    total     = len(rows)
    matched   = sum(1 for r in rows if r.get("Matched") == "✓")
    unmatched = total - matched
    anchors   = {}
    for r in rows:
        at = r.get("Anchor Type") or "Unknown"
        anchors[at] = anchors.get(at, 0) + 1

    summary_data = [
        ("Total Pages Processed", total),
        ("Matched Pairs",         matched),
        ("Unmatched Pages",       unmatched),
        ("", ""),
        ("Breakdown by Anchor", ""),
    ] + [(f"  {k}", v) for k, v in anchors.items()]

    for i, (label, val) in enumerate(summary_data, start=3):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=(val == ""))
        ws2.cell(row=i, column=2, value=val)

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Streamlit UI ──────────────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="EN→ES Activity Card Extractor",
        page_icon="📄",
        layout="wide",
    )

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        st.image(
            "https://img.icons8.com/color/96/000000/pdf.png",
            width=60,
        )
        st.title("Settings")
        font_size_input = st.number_input(
            "Target Font Size (pt)",
            min_value=8, max_value=72,
            value=FONT_SIZE_TARGET,
            help="Spans at this font size (±2 pt) are extracted as primary content."
        )
        tolerance_input = st.number_input(
            "Font Size Tolerance (±pt)",
            min_value=0, max_value=5,
            value=FONT_SIZE_TOLERANCE,
        )
        st.markdown("---")
        st.markdown(
            "**Anchors detected:** ▲ Triangle · ● Circle · ■ Square\n\n"
            "Upload both PDFs and click **Extract**."
        )

    # Use sidebar values as local variables passed to extraction functions
    font_size  = font_size_input
    tolerance  = tolerance_input

    # ── Main ─────────────────────────────────────────────────────────────────
    st.title("📄 EN → ES Activity Card Extractor")
    st.markdown(
        "Upload the **English source PDF** and the **Spanish target PDF**. "
        "The tool extracts font-28 text anchored by ▲ ● ■ symbols "
        "and exports matched source/target pairs to Excel."
    )

    st.info(
        "Large PDFs (500 MB+) are supported. Files are streamed page-by-page "
        "and temp files are deleted immediately after extraction.",
        icon="ℹ️",
    )

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("🇬🇧 Source PDF (English)")
        src_file = st.file_uploader(
            "Upload English PDF", type="pdf", key="src",
            label_visibility="collapsed"
        )
        if src_file:
            st.caption(f"Size: {src_file.size / 1_048_576:.1f} MB")
    with col2:
        st.subheader("🇪🇸 Target PDF (Spanish)")
        tgt_file = st.file_uploader(
            "Upload Spanish PDF", type="pdf", key="tgt",
            label_visibility="collapsed"
        )
        if tgt_file:
            st.caption(f"Size: {tgt_file.size / 1_048_576:.1f} MB")

    run = st.button("🚀 Extract & Generate Excel", use_container_width=True,
                    type="primary")

    if not run:
        st.info("Upload both PDFs and press **Extract** to begin.")
        return

    if not src_file or not tgt_file:
        st.error("Please upload **both** PDFs before extracting.")
        return

    import tempfile, os

    def save_and_extract(uploaded_file, label, font_size, tolerance):
        """Write upload to a temp file then stream-extract to keep RAM low."""
        suffix = ".pdf"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name
            # Stream the upload in 8 MB chunks to avoid one giant .read()
            CHUNK = 8 * 1024 * 1024
            uploaded_file.seek(0)
            while True:
                chunk = uploaded_file.read(CHUNK)
                if not chunk:
                    break
                tmp.write(chunk)

        st.write(f"**{label}** — saved to temp file, extracting…")
        bar = st.progress(0)
        try:
            pages = extract_pages(tmp_path, font_size, tolerance, progress_bar=bar)
        finally:
            os.unlink(tmp_path)   # delete temp file immediately after extraction
        bar.empty()
        return pages

    src_pages = save_and_extract(src_file, "English PDF", font_size, tolerance)
    tgt_pages = save_and_extract(tgt_file, "Spanish PDF", font_size, tolerance)

    with st.spinner("Matching pages & building Excel…"):
        rows       = match_pages(src_pages, tgt_pages)
        excel_data = build_excel(rows)

    # ── Results summary ───────────────────────────────────────────────────────
    total     = len(rows)
    matched   = sum(1 for r in rows if r.get("Matched") == "✓")
    unmatched = total - matched

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Source Pages",    len(src_pages))
    c2.metric("Target Pages",    len(tgt_pages))
    c3.metric("Matched Pairs",   matched)
    c4.metric("Unmatched Pages", unmatched)

    # ── Preview table ─────────────────────────────────────────────────────────
    st.markdown("### Preview")
    preview_cols = [
        "Source Page", "Target Page", "Card ID (Source)",
        "Anchor Symbol", "Anchor Type",
        "Source Title (EN)", "Target Title (ES)", "Matched"
    ]
    import pandas as pd
    df = pd.DataFrame(rows)[preview_cols]
    st.dataframe(df, use_container_width=True, height=380)

    # ── Detailed expander ─────────────────────────────────────────────────────
    with st.expander("🔍 Full extracted data"):
        st.dataframe(pd.DataFrame(rows), use_container_width=True)

    # ── Debug: raw span sizes ─────────────────────────────────────────────────
    with st.expander("🛠 Debug – font sizes detected (first source page)"):
        if src_pages:
            size_map: dict[float, list[str]] = {}
            for sp in src_pages[0]["raw_blocks"]:
                size_map.setdefault(sp["size"], []).append(sp["text"])
            for sz in sorted(size_map.keys(), reverse=True):
                st.write(f"**{sz} pt** → {size_map[sz][:5]}")

    # ── Download ──────────────────────────────────────────────────────────────
    st.download_button(
        label="⬇️ Download Excel Report",
        data=excel_data,
        file_name="EN_ES_Activity_Cards_Extraction.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
